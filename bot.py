import os
import asyncio
import random
import ssl
from logging.handlers import RotatingFileHandler

from email.utils import formatdate
from typing import Optional, Dict, List
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.utils.keyboard import InlineKeyboardBuilder
from email_validator import validate_email, EmailNotValidError
import openpyxl
from openpyxl.workbook import Workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from cryptography.fernet import Fernet
import boto3
from botocore.client import Config
from datetime import datetime
import shutil
import sys
import time
from Yandex_disk import YandexDiskManager
import requests

CURRENT_EDIT_URL = None

import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class FileLock:
    _locked = False
    _last_editor = None

    @classmethod
    def acquire(cls, user_id: int) -> bool:
        if not cls._locked:
            cls._locked = True
            cls._last_editor = user_id
            return True
        return False

    @classmethod
    def release(cls, user_id: int):
        if cls._last_editor == user_id:
            cls._locked = False


async def merge_excel_files(local_path: str, cloud_path: str) -> bool:
    try:
        temp_cloud = "temp_cloud.xlsx"
        await yandex_disk.download_file(cloud_path, temp_cloud)

        wb_local = openpyxl.load_workbook(local_path)
        wb_cloud = openpyxl.load_workbook(temp_cloud)

        for row in wb_cloud.active.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                wb_local.active.append(row)

        wb_local.save(local_path)
        await yandex_disk.upload_file(local_path, cloud_path)
        os.remove(temp_cloud)
        return True
    except Exception as e:
        logger.error(f"Merge error: {e}")
        return False


class ConfigLoader:
    @staticmethod
    def load_key(key_path: str = "secret.key") -> bytes:
        if not os.path.exists(key_path):
            key = Fernet.generate_key()
            with open(key_path, "wb") as f:
                f.write(key)
            logger.info("Генерация нового ключа шифрования")
        return open(key_path, "rb").read()

    @staticmethod
    def encrypt_env_file():
        key = ConfigLoader.load_key()
        cipher = Fernet(key)

        if not os.path.exists(r"C:\Users\kapet\PycharmProjects_bot\pythonProject\random.env"):
            raise FileNotFoundError("Отсутствует .env файл")

        with open(r"C:\Users\kapet\PycharmProjects_bot\pythonProject\random.env", "rb") as f:
            encrypted = cipher.encrypt(f.read())

        with open("config.enc", "wb") as f:
            f.write(encrypted)
        logger.info("Создан новый зашифрованный конфиг")

    @staticmethod
    def decrypt_config() -> Dict[str, str]:
        key = ConfigLoader.load_key()
        cipher = Fernet(key)

        if not os.path.exists("config.enc"):
            if os.path.exists(r"C:\Users\kapet\PycharmProjects_bot\pythonProject\random.env"):
                ConfigLoader.encrypt_env_file()
            else:
                raise FileNotFoundError("Отсутствуют оба файла конфигурации: .env и config.enc")

        with open("config.enc", "rb") as f:
            decrypted = cipher.decrypt(f.read()).decode()

        config_dict = {}
        for line in decrypted.splitlines():
            if "=" in line:
                key, value = line.split("=", 1)
                value = value.split('#')[0].strip()
                config_dict[key.strip()] = value.strip()
        return config_dict


try:
    config = ConfigLoader.decrypt_config()

    if os.path.exists(r"C:\Users\kapet\PycharmProjects_bot\pythonProject\random.env"):
        logger.info("Обнаружен .env файл, обновляю config.enc")
        ConfigLoader.encrypt_env_file()
        config = ConfigLoader.decrypt_config()

    os.environ.update(config)

except Exception as e:
    logger.critical(f"Ошибка загрузки конфига: {str(e)}")
    sys.exit(1)


class AppConfig:
    BOT_TOKEN = os.getenv('BOT_TOKEN')
    ADMINS = [name.strip().lower() for name in os.getenv('ADMINS', '').split(',') if name.strip()]
    NOTIFY_ADMIN = os.getenv('NOTIFY_ADMIN', '').strip()

    YC_ACCESS_KEY = os.getenv('YC_ACCESS_KEY')
    YC_SECRET_KEY = os.getenv('YC_SECRET_KEY')
    YC_BUCKET_NAME = os.getenv('YC_BUCKET_NAME')

    SMTP_SERVER = os.getenv('SMTP_SERVER')
    SMTP_PORT = int(os.getenv('SMTP_PORT', '587').split()[0])
    EMAIL_LOGIN = os.getenv('EMAIL_LOGIN')
    EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD')
    EMAIL_FROM = os.getenv('EMAIL_FROM')

    DELAY_TG = float(os.getenv('DELAY_TG', 1))
    BACKUP_INTERVAL = int(os.getenv('BACKUP_INTERVAL', 3600))
    SECRET_CODE = os.getenv('SECRET_CODE', 'DEFAULT_CODE')
    VERIFICATION_STRING = os.getenv('VERIFICATION_STRING', 'DEFAULT_VERIFICATION')
    DATA_HASH_SALT = os.getenv('DATA_HASH_SALT', 'DEFAULT_SALT')

    AIRTABLE_API_KEY = os.getenv('AIRTABLE_API_KEY')
    AIRTABLE_BASE_ID = os.getenv('AIRTABLE_BASE_ID')


if not all([AppConfig.BOT_TOKEN, AppConfig.ADMINS]):
    print("Не загружены обязательные переменные окружения!")
    sys.exit(1)

storage = MemoryStorage()
bot = Bot(token=AppConfig.BOT_TOKEN)
dp = Dispatcher(storage=storage)

for directory in ["backups", "logs", "temp", "enc"]:
    os.makedirs(directory, exist_ok=True)


def setup_logging():
    log_formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s'
    )

    file_handler = RotatingFileHandler(
        'logs/bot.log',
        maxBytes=5 * 1024 * 1024,
        backupCount=3,
        encoding='utf-8'
    )
    file_handler.setFormatter(log_formatter)

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(log_formatter)

    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logging()




class Form(StatesGroup):
    fio = State()
    email = State()
    code = State()


class AdminForm(StatesGroup):
    confirm_action = State()
    secret_code = State()
    results_file = State()


class CloudStorage:
    def __init__(self):
        self.s3 = boto3.client(
            's3',
            endpoint_url='https://storage.yandexcloud.net',
            aws_access_key_id=AppConfig.YC_ACCESS_KEY,
            aws_secret_access_key=AppConfig.YC_SECRET_KEY,
            config=Config(signature_version='s3v4'))

        try:
            self.s3.head_bucket(Bucket=AppConfig.YC_BUCKET_NAME)
            logger.info("Cloud storage connected successfully")
        except Exception as e:
            logger.critical(f"Cloud storage connection error: {e}")
            sys.exit(1)

    async def upload_file(self, local_path: str, cloud_path: str) -> bool:
        try:
            self.s3.upload_file(local_path, AppConfig.YC_BUCKET_NAME, cloud_path)
            logger.info(f"Uploaded {local_path} to {cloud_path}")
            return True
        except Exception as e:
            logger.error(f"Upload error: {e}")
            return False

    async def download_file(self, cloud_path: str, local_path: str) -> bool:
        try:
            self.s3.download_file(AppConfig.YC_BUCKET_NAME, cloud_path, local_path)
            logger.info(f"Downloaded {cloud_path} to {local_path}")
            return True
        except Exception as e:
            logger.error(f"Download error: {e}")
            return False

    async def get_file_url(self, cloud_path: str, expires_in: int = 3600) -> str:
        try:
            url = self.s3.generate_presigned_url(
                'get_object',
                Params={'Bucket': AppConfig.YC_BUCKET_NAME, 'Key': cloud_path},
                ExpiresIn=expires_in
            )
            return url
        except Exception as e:
            logger.error(f"URL generation error: {e}")
            return ""

    async def set_public_access(self, file_key: str):
        try:
            self.s3.put_object_acl(
                Bucket=AppConfig.YC_BUCKET_NAME,
                Key=file_key,
                ACL='public-read'
            )
            logger.info(f"Public access granted for {file_key}")
            return True
        except Exception as e:
            logger.error(f"ACL error: {e}")
            return False

    async def set_private_access(self, file_key: str):
        try:
            self.s3.put_object_acl(
                Bucket=AppConfig.YC_BUCKET_NAME,
                Key=file_key,
                ACL='private'
            )
            logger.info(f"Private access set for {file_key}")
            return True
        except Exception as e:
            logger.error(f"ACL error: {e}")
            return False


storage = CloudStorage()


def is_admin(user: types.User) -> bool:
    user_identifiers = {
        str(user.id).lower(),
        f"@{user.username.lower()}" if user.username else None,
        user.username.lower() if user.username else None
    }
    user_identifiers.discard(None)

    admins = [a.strip().lower() for a in AppConfig.ADMINS if a.strip()]

    logger.info(f"Checking admin access for: {user_identifiers}")
    logger.info(f"Admin list: {admins}")

    is_adm = any(admin in user_identifiers for admin in admins)
    logger.info(f"Access {'granted' if is_adm else 'denied'}")
    return is_adm


async def user_exists(user_id: int, username: Optional[str] = None, fio: Optional[str] = None) -> bool:
    try:
        wb = await asyncio.to_thread(openpyxl.load_workbook, "master_data.xlsx")
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if (str(user_id) == row[2]
                    or (username and f"@{username.lower()}" == row[2].lower())
                    or (fio and fio.lower() == row[0].lower())):
                return True
        return False
    except Exception as e:
        logger.error(f"User check error: {e}")
        return False


async def add_user(fio: str, email: str, tg_username: str) -> bool:
    try:
        wb = await asyncio.to_thread(openpyxl.load_workbook, "master_data.xlsx")
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if fio.lower() == row[0].lower() or email.lower() == row[1].lower():
                return False

        ws.append([fio, email, f"@{tg_username}" if tg_username else "", 0])

        await asyncio.to_thread(wb.save, "master_data.xlsx")
        await storage.upload_file("master_data.xlsx", "master_data.xlsx")
        return True

    except Exception as e:
        logger.error(f"Add user error: {e}")
        return False


async def create_encrypted_version(source_path: str, dest_path: str) -> bool:
    try:
        key = ConfigLoader.load_key()
        cipher = Fernet(key)

        with open(source_path, "rb") as f:
            encrypted = cipher.encrypt(f.read())

        with open(dest_path, "wb") as f:
            f.write(encrypted)

        logger.info(f"Created encrypted version: {dest_path}")
        return True
    except Exception as e:
        logger.error(f"Encryption error: {e}")
        return False


async def restore_from_encrypted(enc_path: str, dest_path: str) -> bool:
    try:
        key = ConfigLoader.load_key()
        cipher = Fernet(key)

        with open(enc_path, "rb") as f:
            decrypted = cipher.decrypt(f.read())

        with open(dest_path, "wb") as f:
            f.write(decrypted)

        logger.info(f"Restored from encrypted: {dest_path}")
        return True
    except Exception as e:
        logger.error(f"Decryption error: {e}")
        return False


async def send_email(to: str, subject: str, body: str) -> bool:
    try:
        if not all([AppConfig.SMTP_SERVER, AppConfig.SMTP_PORT,
                    AppConfig.EMAIL_LOGIN, AppConfig.EMAIL_PASSWORD]):
            logger.error("Не все SMTP параметры настроены!")
            return False

        msg = MIMEMultipart()
        msg['From'] = AppConfig.EMAIL_FROM
        msg['To'] = to
        msg['Subject'] = subject
        msg['Date'] = formatdate(localtime=True)
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        context = ssl.create_default_context()

        with smtplib.SMTP_SSL(
                host=AppConfig.SMTP_SERVER,
                port=AppConfig.SMTP_PORT,
                context=context,
                timeout=10
        ) as server:
            server.login(AppConfig.EMAIL_LOGIN, AppConfig.EMAIL_PASSWORD)
            server.send_message(msg)

        logger.info(f"Email отправлен на {to}")
        return True

    except smtplib.SMTPAuthenticationError:
        logger.error("Ошибка аутентификации: неверный логин/пароль")
    except Exception as e:
        logger.error(f"Ошибка отправки: {str(e)}")

    return False


async def validate_email_address(email: str) -> bool:
    try:
        validate_email(email)
        return True
    except EmailNotValidError:
        return False


async def generate_verification_code(email: str) -> Optional[str]:
    try:
        code = str(random.randint(100000, 999999))
        subject = "Код подтверждения"
        body = f"Ваш код подтверждения: {code}"

        if await send_email(email, subject, body):
            return code
        return None

    except Exception as e:
        logger.error(f"Ошибка генерации кода: {str(e)}")
        return None


async def is_valid_excel(file_path: str, required_headers: Optional[List[str]] = None) -> bool:

    if required_headers is None:
        required_headers = ["ФИО", "Email", "Telegram", "Статус", "Проверка"]

    try:
        wb = await asyncio.to_thread(openpyxl.load_workbook, file_path)
        ws = wb.active
        actual_headers = [cell.value for cell in ws[1]]
        return actual_headers == required_headers
    except Exception as e:
        logger.error(f"Excel validation error: {e}")
        return False


async def init_master_data() -> bool:
    try:
        if not os.path.exists("master_data.xlsx"):
            wb = Workbook()
            ws = wb.active
            ws.append(["ФИО", "Email", "Telegram", "Статус", "Проверка"])
            wb.save("master_data.xlsx")
            await storage.upload_file("master_data.xlsx", "master_data.xlsx")

            await create_encrypted_version("master_data.xlsx", "enc/master_data.enc")

            logger.info("Created new master_data.xlsx with encrypted version")
        elif not await is_valid_excel("master_data.xlsx"):
            logger.error("Invalid master_data.xlsx structure! Creating backup...")
            backup_name = f"backups/corrupted_master_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            shutil.move("master_data.xlsx", backup_name)

            if os.path.exists("enc/master_data.enc"):
                if await restore_from_encrypted("enc/master_data.enc", "master_data.xlsx"):
                    if await is_valid_excel("master_data.xlsx"):
                        logger.info("Restored from encrypted version successfully")
                        return True

            return await init_master_data()
        return True
    except Exception as e:
        logger.error(f"Init error: {e}")
        return False


async def backup_data() -> bool:
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if os.path.exists("master_data.xlsx"):
            backup_name = f"backups/master_data_{timestamp}.xlsx"
            shutil.copy2("master_data.xlsx", backup_name)
            await storage.upload_file(backup_name, f"backups/{os.path.basename(backup_name)}")

            enc_backup = f"backups/master_data_{timestamp}.enc"
            if os.path.exists("enc/master_data.enc"):
                shutil.copy2("enc/master_data.enc", enc_backup)
                await storage.upload_file(enc_backup, f"backups/{os.path.basename(enc_backup)}")

        if os.path.exists("logs/bot.log"):
            log_backup = f"backups/logs_{timestamp}.log"
            shutil.copy2("logs/bot.log", log_backup)

        backups = sorted([f for f in os.listdir("backups") if f.startswith(("master_data_", "logs_"))])
        for old_backup in backups[:-10]:
            os.remove(f"backups/{old_backup}")

        logger.info("Backup completed with encrypted versions")
        return True
    except Exception as e:
        logger.error(f"Backup error: {e}")
        return False


async def send_template_file(chat_id: int):
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(["ФИО", "Email", "Telegram", "Статус", "Проверка"])
        template_path = "temp/template.xlsx"
        wb.save(template_path)

        with open(template_path, "rb") as f:
            await bot.send_document(chat_id, types.BufferedInputFile(f.read(), filename="template.xlsx"))

        os.remove(template_path)
    except Exception as e:
        logger.error(f"Template send error: {e}")


def get_user_keyboard() -> types.InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.button(text="📝 Регистрация", callback_data="start_registration")
    builder.button(text="ℹ️ Помощь", callback_data="user_help")
    return builder.as_markup()


def get_admin_keyboard() -> types.InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.button(
        text="✏️ Онлайн-редактирование",
        callback_data="admin_online_edit"
    )
    builder.button(
        text="💾 Сохранить изменения",
        callback_data="save_online_edit"
    )
    builder.button(
        text="❌ Отменить редактирование",
        callback_data="cancel_online_edit"
    )
    builder.button(
        text="📨 Рассылка результатов",
        callback_data="admin_mailing"
    )
    builder.adjust(1)
    return builder.as_markup()


def get_confirmation_keyboard() -> types.InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.button(text="✅ Подтвердить", callback_data="confirm_yes")
    builder.button(text="❌ Отменить", callback_data="confirm_no")
    return builder.as_markup()


@dp.message(Command("start"))
async def start_handler(message: types.Message):
    if is_admin(message.from_user):
        await message.answer("Админ-панель:", reply_markup=get_admin_keyboard())
    else:
        await message.answer("Добро пожаловать!", reply_markup=get_user_keyboard())


@dp.message(Command("help"))
async def help_handler(message: types.Message):
    await message.answer(
        "Доступные команды:\n"
        "/start - Главное меню\n"
        "/help - Справка\n\n"
        "Для регистрации нажмите кнопку ниже:",
        reply_markup=get_user_keyboard()
    )


yandex_disk = YandexDiskManager()


@dp.callback_query(F.data == "admin_online_edit")
async def admin_online_edit(callback: types.CallbackQuery):
    """Обработчик кнопки онлайн-редактирования"""
    if not is_admin(callback.from_user):
        await callback.answer("Доступ запрещён!")
        return

    try:
        if not await yandex_disk.upload_file("master_data.xlsx", "/master_data.xlsx"):
            raise Exception("Не удалось загрузить файл на Яндекс.Диск")

        edit_url = await yandex_disk.get_edit_url("/master_data.xlsx")
        if not edit_url:
            raise Exception("Не удалось получить ссылку для редактирования")

        await callback.message.answer(
            "✏️ Редактируйте файл по ссылке:\n"
            f"{edit_url}\n\n"
            "После завершения нажмите 'Сохранить изменения' в меню",
            reply_markup=get_admin_keyboard()
        )
    except Exception as e:
        await callback.message.answer(f"❌ Ошибка: {str(e)}")
    finally:
        await callback.answer()

@dp.callback_query(F.data == "save_online_edit")
async def save_online_edit(callback: types.CallbackQuery):
    """Сохранение изменений из онлайн-редактора"""
    if not is_admin(callback.from_user):
        await callback.answer("Доступ запрещён!")
        return

    try:
        # Скачиваем обновленную версию
        if not await yandex_disk.download_file("/master_data.xlsx", "master_data.xlsx"):
            raise Exception("Не удалось скачать обновленный файл")

        await callback.answer("✅ Изменения сохранены")
        await callback.message.answer(
            "Файл успешно обновлен!",
            reply_markup=get_admin_keyboard()
        )
    except Exception as e:
        await callback.message.answer(f"❌ Ошибка сохранения: {str(e)}")
    finally:
        await callback.answer()

@dp.callback_query(F.data == "cancel_online_edit")
async def cancel_online_edit(callback: types.CallbackQuery):
    if not is_admin(callback.from_user):
        await callback.answer("Доступ запрещён!")
        return

    try:
        await yandex_disk.set_private_access("/master_data.xlsx")
        await callback.answer("🔒 Редактирование отменено")
        await callback.message.answer(
            "Доступ к редактированию закрыт",
            reply_markup=get_admin_keyboard()
        )
    except Exception as e:
        await callback.message.answer(f"❌ Ошибка: {str(e)}")
    finally:
        await callback.answer()




@dp.callback_query(F.data == "admin_mailing")
async def start_mailing(callback: types.CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user):
        await callback.answer("Доступ запрещён!")
        return

    await callback.message.answer("Для начала рассылки введите секретный код:")
    await state.set_state(AdminForm.secret_code)
    await callback.answer()



@dp.callback_query(F.data == "start_registration")
async def start_registration(callback: types.CallbackQuery, state: FSMContext):
    if await user_exists(callback.from_user.id, callback.from_user.username):
        await callback.answer("Вы уже зарегистрированы!")
        return

    await callback.message.answer("Введите ваше ФИО (только буквы и пробелы):")
    await state.set_state(Form.fio)
    await callback.answer()


@dp.message(Form.fio)
async def process_fio(message: types.Message, state: FSMContext):
    if not message.text.replace(" ", "").isalpha():
        await message.answer("ФИО должно содержать только буквы и пробелы. Введите заново:")
        return

    if await user_exists(None, None, message.text):
        await message.answer("Пользователь с таким ФИО уже зарегистрирован!")
        await state.clear()
        return

    await state.update_data(fio=message.text)
    await message.answer("Теперь введите ваш email:")
    await state.set_state(Form.email)


@dp.message(Form.email)
async def process_email(message: types.Message, state: FSMContext):
    if not await validate_email_address(message.text):
        await message.answer("❌ Неверный формат email. Введите заново:")
        return

    email = message.text
    await state.update_data(email=email)

    builder = InlineKeyboardBuilder()
    builder.button(text="🔄 Отправить код повторно", callback_data="resend_code")

    try:
        code = await generate_verification_code(email)
        if not code:
            raise Exception("Не удалось сгенерировать код")

        await state.update_data(code=code)
        await message.answer(
            "📨 Код подтверждения отправлен на вашу почту.\n"
            "Проверьте папку 'Спам', если письмо не пришло.\n\n"
            "Если код не приходит, нажмите кнопку ниже:",
            reply_markup=builder.as_markup()
        )
        await state.set_state(Form.code)
    except Exception as e:
        logger.error(f"Ошибка верификации: {str(e)}")
        await message.answer(
            "⚠️ Не удалось отправить код подтверждения.\n"
            "Попробуйте позже или обратитесь к администратору."
        )
        await state.clear()


@dp.callback_query(F.data == "resend_code", Form.code)
async def resend_code_handler(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    email = data.get("email")

    builder = InlineKeyboardBuilder()
    builder.button(text="🔄 Попробовать ещё раз", callback_data="resend_code")
    builder.button(text="✉️ Связаться с администратором", url=f"tg://user?id={AppConfig.ADMINS}")

    try:
        new_code = str(random.randint(100000, 999999))
        await state.update_data(code=new_code)

        if await send_email(email, "Код подтверждения", f"Ваш новый код: {new_code}"):
            await callback.message.edit_text(
                "📨 Новый код подтверждения отправлен.\n"
                "Если письмо не приходит:\n"
                "1. Проверьте папку 'Спам'\n"
                "2. Убедитесь в правильности email\n"
                "3. Обратитесь к администратору",
                reply_markup=builder.as_markup()
            )
        else:
            raise Exception("Ошибка отправки email")
    except Exception as e:
        await callback.message.edit_text(
            "❌ Не удалось отправить код.\n"
            "Пожалуйста, обратитесь к администратору:",
            reply_markup=builder.as_markup()
        )
    finally:
        await callback.answer()


@dp.message(Command("myinfo"))
async def myinfo(message: types.Message):
    info = (
        f"Ваш ID: {message.from_user.id}\n"
        f"Username: @{message.from_user.username}\n"
        f"Администратор: {'Да' if is_admin(message.from_user) else 'Нет'}\n"
        f"Список админов: {AppConfig.ADMINS}"
    )
    await message.answer(info)


@dp.message(Form.code)
async def process_code(message: types.Message, state: FSMContext):
    data = await state.get_data()
    if message.text != data["code"]:
        await message.answer("Неверный код. Попробуйте снова:")
        return

    tg_username = message.from_user.username
    if await add_user(data["fio"], data["email"], tg_username):
        await message.answer("✅ Регистрация завершена!")
    else:
        await message.answer("❌ Ошибка сохранения данных. Попробуйте позже.")
    await state.clear()





@dp.message(AdminForm.secret_code)
async def check_secret_code(message: types.Message, state: FSMContext):
    if message.text != AppConfig.SECRET_CODE:
        await message.answer("Неверный код! Рассылка отменена.")
        await state.clear()
        return

    await message.answer("Код верный. Отправьте файл results.xlsx:")
    await state.set_state(AdminForm.results_file)


@dp.message(F.document & F.document.file_name.endswith('.xlsx'), AdminForm.results_file)
async def process_mailing_file(message: types.Message, state: FSMContext):
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"temp/results_{timestamp}.xlsx"

        await bot.download(message.document, destination=file_path)

        if not await is_valid_excel(file_path, ["ФИО", "Email", "Результат"]):
            raise ValueError("Файл должен содержать колонки: ФИО, Email, Результат")

        await storage.upload_file(file_path, f"results/results_{timestamp}.xlsx")
        await message.answer("⏳ Начинаю рассылку результатов...")

        wb_main = await asyncio.to_thread(openpyxl.load_workbook, "master_data.xlsx")
        ws_main = wb_main.active

        wb_results = await asyncio.to_thread(openpyxl.load_workbook, file_path)
        ws_results = wb_results.active

        success = 0
        errors = 0

        for row in ws_results.iter_rows(min_row=2, values_only=True):
            fio, email, result = row[0], row[1], row[2]
            user_found = False

            for idx, user_row in enumerate(ws_main.iter_rows(min_row=2, values_only=True), start=2):
                if user_row[0] == fio:
                    email_sent = await send_email(
                        email,
                        "Результаты",
                        f"Уважаемый(ая) {fio}!\nВаш результат: {result}"
                    )

                    tg_username = user_row[2].replace('@', '') if '@' in user_row[2] else user_row[2]
                    tg_sent = await bot.send_message(
                        tg_username,
                        f"Ваш результат: {result}"
                    ) if tg_username else False

                    status = 3 if email_sent and tg_sent else \
                        2 if email_sent else \
                            1 if tg_sent else 4

                    ws_main.cell(row=idx, column=4, value=status)
                    user_found = True

                    if status != 4:
                        success += 1
                    else:
                        errors += 1
                    break

            if not user_found:
                logger.error(f"Пользователь не найден: {fio}")
                errors += 1

            await asyncio.sleep(AppConfig.DELAY_TG)

        await asyncio.to_thread(wb_main.save, "master_data.xlsx")
        await storage.upload_file("master_data.xlsx", "master_data.xlsx")

        await message.answer(
            f"📊 Рассылка завершена:\n"
            f"• Успешно: {success}\n"
            f"• Ошибки: {errors}"
        )
        logger.info(f"Mailing completed. Success: {success}, Errors: {errors}")

    except Exception as e:
        logger.error(f"Mailing error: {e}")
        await message.answer(f"❌ Ошибка рассылки: {str(e)}")
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)
        await state.clear()


async def periodic_backup():
    while True:
        await asyncio.sleep(AppConfig.BACKUP_INTERVAL)
        try:
            await backup_data()
        except Exception as e:
            logger.error(f"Periodic backup error: {e}")


async def auto_sync_loop():
    """Фоновая задача для автосинхронизации"""
    while True:
        try:
            cloud_ver = await yandex_disk.get_file_version("/master_data.xlsx")
            local_ver = datetime.fromtimestamp(os.path.getmtime("master_data.xlsx")).isoformat()

            # Добавляем проверку на None
            if cloud_ver and local_ver and cloud_ver > local_ver:
                await merge_excel_files("master_data.xlsx", "/master_data.xlsx")
                logger.info("Автосинхронизация выполнена")
        except Exception as e:
            logger.error(f"Auto-sync error: {e}")

        await asyncio.sleep(5)

async def on_startup():
    logger.info("=" * 50)
    logger.info("Запуск бота. Проверка окружения...")

    required_vars = ['BOT_TOKEN', 'YANDEX_DISK_TOKEN', 'ADMINS']
    for var in required_vars:
        if not os.getenv(var):
            logger.critical(f"Отсутствует переменная окружения: {var}")

    if not os.path.exists("master_data.xlsx"):
        logger.warning("Файл master_data.xlsx не найден, будет создан новый")

    logger.info("Проверка Яндекс.Диска...")
    try:
        await yandex_disk.upload_file("test.txt", "/test.txt")
    except Exception as e:
        logger.critical(f"Ошибка подключения к Яндекс.Диску: {str(e)}")

    logger.info("Запуск завершен успешно")
    logger.info("=" * 50)

    test_email = "kapetanamerika77@gmail.com"

    logger.info("Проверка SMTP соединения...")
    smtp_ok = await send_email(
        to=test_email,
        subject="Тест SMTP соединения",
        body="Это тестовое сообщение при запуске бота"
    )

    if smtp_ok:
        logger.info("SMTP соединение успешно проверено")
    else:
        logger.critical("SMTP соединение не работает! Проверьте настройки")

    logger.info("Starting bot...")
    await init_master_data()
    asyncio.create_task(periodic_backup())

    if AppConfig.NOTIFY_ADMIN and AppConfig.NOTIFY_ADMIN.strip():
        try:
            if AppConfig.NOTIFY_ADMIN.isdigit():
                user_id = int(AppConfig.NOTIFY_ADMIN)
                await bot.send_message(user_id, "🤖 Бот успешно запущен")

            elif AppConfig.NOTIFY_ADMIN.startswith('@'):
                username = AppConfig.NOTIFY_ADMIN
                user = await bot.get_chat(username)
                await bot.send_message(user.id, "🤖 Бот успешно запущен")

            else:
                logger.warning(f"Некорректный формат NOTIFY_ADMIN: {AppConfig.NOTIFY_ADMIN}")

        except Exception as e:
            logger.error(f"Не удалось отправить уведомление админу: {str(e)}")
    else:
        logger.info("NOTIFY_ADMIN не настроен, пропускаем уведомление")
    asyncio.create_task(auto_sync_loop())


async def on_shutdown():
    logger.info("Stopping bot...")
    await backup_data()

    if AppConfig.NOTIFY_ADMIN and AppConfig.NOTIFY_ADMIN.strip():
        try:
            if AppConfig.NOTIFY_ADMIN.isdigit() or AppConfig.NOTIFY_ADMIN.startswith('@'):
                await bot.send_message(AppConfig.NOTIFY_ADMIN, "🛑 Бот остановлен")
        except Exception as e:
            logger.error(f"Ошибка при отправке уведомления: {str(e)}")


async def main():
    dp.startup.register(on_startup)
    dp.shutdown.register(on_shutdown)

    try:
        await dp.start_polling(bot)
    finally:
        await bot.session.close()


if __name__ == "__main__":
    asyncio.run(main())
